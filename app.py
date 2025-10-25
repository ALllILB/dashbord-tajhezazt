from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash
from functools import wraps
import bcrypt
import re
from database import get_db, init_db, validate_ip
from werkzeug.utils import secure_filename
import openpyxl
import io
import os

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'
init_db()

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
            user = cursor.fetchone()
            
            if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['role'] = user['role']
                flash('Login successful!', 'success')
                return redirect(url_for('index'))
            else:
                flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
@admin_required
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        
        if not username or not password:
            flash('Username and password are required', 'error')
            return render_template('register.html')
        
        with get_db() as conn:
            cursor = conn.cursor()
            # Check if user exists
            cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
            if cursor.fetchone():
                flash('Username already exists', 'error')
                return render_template('register.html')
            
            # Create new user
            hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
            cursor.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                         (username, hashed_password.decode('utf-8'), role))
            conn.commit()
            flash('User created successfully', 'success')
            return redirect(url_for('index'))
    
    return render_template('register.html')

@app.route('/equipment-dashboard')
@login_required
def equipment_dashboard():
    return render_template('equipment-dashboard.html')

@app.route('/infrastructure-dashboard')
@login_required
def infrastructure_dashboard():
    return render_template('infrastructure-dashboard.html')

# API Routes for systems
@app.route('/api/systems')
@login_required
def get_systems():
    location = request.args.get('location')
    with get_db() as conn:
        cursor = conn.cursor()
        if location:
            cursor.execute('SELECT * FROM systems WHERE location = ?', (location,))
        else:
            cursor.execute('SELECT * FROM systems')
        systems = [dict(row) for row in cursor.fetchall()]
        return jsonify(systems)

@app.route('/api/systems/<int:id>', methods=['PUT', 'DELETE'])
@admin_required
def manage_system(id):
    with get_db() as conn:
        cursor = conn.cursor()
        
        if request.method == 'PUT':
            data = request.get_json()
            if not validate_ip(data['ip_address']):
                return jsonify({'error': 'Invalid IP address'}), 400
            
            cursor.execute('''
                UPDATE systems 
                SET location=?, system_name=?, user=?, ip_address=?, 
                    antivirus_status=?, firewall_status=?
                WHERE id=?
            ''', (data['location'], data['system_name'], data['user'], 
                  data['ip_address'], data['antivirus_status'], 
                  data['firewall_status'], id))
            conn.commit()
            return jsonify({'success': True})
        
        elif request.method == 'DELETE':
            cursor.execute('DELETE FROM systems WHERE id=?', (id,))
            conn.commit()
            return jsonify({'success': True})

@app.route('/api/systems', methods=['POST'])
@admin_required
def add_system():
    data = request.get_json()
    if not validate_ip(data['ip_address']):
        return jsonify({'error': 'Invalid IP address'}), 400
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO systems (location, system_name, user, ip_address, 
                                antivirus_status, firewall_status)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (data['location'], data['system_name'], data['user'], 
              data['ip_address'], data['antivirus_status'], 
              data['firewall_status']))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})

# API Routes for telephony
@app.route('/api/telephony')
@login_required
def get_telephony():
    location = request.args.get('location')
    with get_db() as conn:
        cursor = conn.cursor()
        if location:
            cursor.execute('SELECT * FROM telephony WHERE location = ?', (location,))
        else:
            cursor.execute('SELECT * FROM telephony')
        telephony = [dict(row) for row in cursor.fetchall()]
        return jsonify(telephony)

@app.route('/api/telephony/<int:id>', methods=['PUT', 'DELETE'])
@admin_required
def manage_telephony(id):
    with get_db() as conn:
        cursor = conn.cursor()
        
        if request.method == 'PUT':
            data = request.get_json()
            cursor.execute('''
                UPDATE telephony 
                SET location=?, personnel_name=?, internal_number=?, 
                    phone_type=?, upgrade_needed=?
                WHERE id=?
            ''', (data['location'], data['personnel_name'], data['internal_number'], 
                  data['phone_type'], data['upgrade_needed'], id))
            conn.commit()
            return jsonify({'success': True})
        
        elif request.method == 'DELETE':
            cursor.execute('DELETE FROM telephony WHERE id=?', (id,))
            conn.commit()
            return jsonify({'success': True})

@app.route('/api/telephony', methods=['POST'])
@admin_required
def add_telephony():
    data = request.get_json()
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO telephony (location, personnel_name, internal_number, 
                                  phone_type, upgrade_needed)
            VALUES (?, ?, ?, ?, ?)
        ''', (data['location'], data['personnel_name'], data['internal_number'], 
              data['phone_type'], data['upgrade_needed']))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})

# API Routes for CCTV
@app.route('/api/cctv')
@login_required
def get_cctv():
    location = request.args.get('location')
    with get_db() as conn:
        cursor = conn.cursor()
        if location:
            cursor.execute('SELECT * FROM cctv WHERE location = ?', (location,))
        else:
            cursor.execute('SELECT * FROM cctv')
        cctv = [dict(row) for row in cursor.fetchall()]
        return jsonify(cctv)

@app.route('/api/cctv/<int:id>', methods=['PUT', 'DELETE'])
@admin_required
def manage_cctv(id):
    with get_db() as conn:
        cursor = conn.cursor()
        
        if request.method == 'PUT':
            data = request.get_json()
            cursor.execute('''
                UPDATE cctv 
                SET location=?, point_needed=?, priority=?, reason=?
                WHERE id=?
            ''', (data['location'], data['point_needed'], data['priority'], 
                  data['reason'], id))
            conn.commit()
            return jsonify({'success': True})
        
        elif request.method == 'DELETE':
            cursor.execute('DELETE FROM cctv WHERE id=?', (id,))
            conn.commit()
            return jsonify({'success': True})

@app.route('/api/cctv', methods=['POST'])
@admin_required
def add_cctv():
    data = request.get_json()
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO cctv (location, point_needed, priority, reason)
            VALUES (?, ?, ?, ?)
        ''', (data['location'], data['point_needed'], data['priority'], 
              data['reason']))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})

# API Routes for infrastructure
@app.route('/api/infrastructure')
@login_required
def get_infrastructure():
    location = request.args.get('location')
    with get_db() as conn:
        cursor = conn.cursor()
        if location:
            cursor.execute('SELECT * FROM infrastructure WHERE location = ?', (location,))
        else:
            cursor.execute('SELECT * FROM infrastructure')
        infrastructure = [dict(row) for row in cursor.fetchall()]
        return jsonify(infrastructure)

@app.route('/api/infrastructure/<int:id>', methods=['PUT', 'DELETE'])
@admin_required
def manage_infrastructure(id):
    with get_db() as conn:
        cursor = conn.cursor()
        
        if request.method == 'PUT':
            data = request.get_json()
            cursor.execute('''
                UPDATE infrastructure 
                SET location=?, printer=?, computer=?, voip_phone=?, camera=?,
                    attendance_clock=?, nutrition_clock=?, nvr=?, managed_switch=?,
                    unmanaged_switch=?, server=?, voip_pbx=?, firewall=?
                WHERE id=?
            ''', (data['location'], data['printer'], data['computer'], 
                  data['voip_phone'], data['camera'], data['attendance_clock'],
                  data['nutrition_clock'], data['nvr'], data['managed_switch'],
                  data['unmanaged_switch'], data['server'], data['voip_pbx'],
                  data['firewall'], id))
            conn.commit()
            return jsonify({'success': True})
        
        elif request.method == 'DELETE':
            cursor.execute('DELETE FROM infrastructure WHERE id=?', (id,))
            conn.commit()
            return jsonify({'success': True})

@app.route('/api/infrastructure', methods=['POST'])
@admin_required
def add_infrastructure():
    data = request.get_json()
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO infrastructure (location, printer, computer, voip_phone, camera,
                                       attendance_clock, nutrition_clock, nvr, managed_switch,
                                       unmanaged_switch, server, voip_pbx, firewall)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (data['location'], data['printer'], data['computer'], 
              data['voip_phone'], data['camera'], data['attendance_clock'],
              data['nutrition_clock'], data['nvr'], data['managed_switch'],
              data['unmanaged_switch'], data['server'], data['voip_pbx'],
              data['firewall']))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})

# API Routes for user management
@app.route('/api/users')
@admin_required
def get_users():
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id, username, role FROM users')
        users = [dict(row) for row in cursor.fetchall()]
        return jsonify(users)

@app.route('/api/users/<int:id>', methods=['PUT', 'DELETE'])
@admin_required
def manage_user(id):
    with get_db() as conn:
        cursor = conn.cursor()
        
        if request.method == 'PUT':
            data = request.get_json()
            
            # If password is provided, hash it
            if data.get('password'):
                hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt())
                cursor.execute('UPDATE users SET username=?, password=?, role=? WHERE id=?',
                             (data['username'], hashed_password.decode('utf-8'), data['role'], id))
            else:
                cursor.execute('UPDATE users SET username=?, role=? WHERE id=?',
                             (data['username'], data['role'], id))
            
            conn.commit()
            return jsonify({'success': True})
        
        elif request.method == 'DELETE':
            # Prevent deletion of the current user
            if id == session['user_id']:
                return jsonify({'error': 'Cannot delete your own account'}), 400
            
            cursor.execute('DELETE FROM users WHERE id=?', (id,))
            conn.commit()
            return jsonify({'success': True})

@app.route('/api/users', methods=['POST'])
@admin_required
def add_user():
    data = request.get_json()
    
    if not data.get('username') or not data.get('password'):
        return jsonify({'error': 'Username and password are required'}), 400
    
    with get_db() as conn:
        cursor = conn.cursor()
        
        # Check if user exists
        cursor.execute('SELECT id FROM users WHERE username = ?', (data['username'],))
        if cursor.fetchone():
            return jsonify({'error': 'Username already exists'}), 400
        
        # Create new user
        hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt())
        cursor.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                     (data['username'], hashed_password.decode('utf-8'), data['role']))
        conn.commit()
        return jsonify({'success': True, 'id': cursor.lastrowid})

# Import/Export routes
@app.route('/api/import/<table>', methods=['POST'])
@admin_required
def import_data(table):
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and file.filename.endswith('.xlsx'):
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file.read()))
            sheet = workbook.active
            
            with get_db() as conn:
                cursor = conn.cursor()
                
                if table == 'systems':
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:  # Skip empty rows
                            cursor.execute('''
                                INSERT INTO systems (location, system_name, user, ip_address, 
                                                   antivirus_status, firewall_status)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', row)
                
                elif table == 'telephony':
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            cursor.execute('''
                                INSERT INTO telephony (location, personnel_name, internal_number, 
                                                     phone_type, upgrade_needed)
                                VALUES (?, ?, ?, ?, ?)
                            ''', row)
                
                elif table == 'cctv':
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            cursor.execute('''
                                INSERT INTO cctv (location, point_needed, priority, reason)
                                VALUES (?, ?, ?, ?)
                            ''', row)
                
                elif table == 'infrastructure':
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            cursor.execute('''
                                INSERT INTO infrastructure (location, printer, computer, voip_phone, 
                                                          camera, attendance_clock, nutrition_clock, 
                                                          nvr, managed_switch, unmanaged_switch, 
                                                          server, voip_pbx, firewall)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', row)
                
                conn.commit()
                return jsonify({'success': True, 'imported': sheet.max_row - 1})
        
        except Exception as e:
            return jsonify({'error': str(e)}), 400
    
    return jsonify({'error': 'Invalid file format'}), 400

# Admin management pages
@app.route('/admin/systems')
@admin_required
def admin_systems():
    return render_template('admin/systems.html')

@app.route('/admin/telephony')
@admin_required
def admin_telephony():
    return render_template('admin/telephony.html')

@app.route('/admin/cctv')
@admin_required
def admin_cctv():
    return render_template('admin/cctv.html')

@app.route('/admin/infrastructure')
@admin_required
def admin_infrastructure():
    return render_template('admin/infrastructure.html')

@app.route('/admin/users')
@admin_required
def admin_users():
    return render_template('admin/users.html')

if __name__ == '__main__':
    app.run(debug=True)